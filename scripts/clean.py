import os, re, glob, sys

WORK_DIR = os.path.dirname(os.path.abspath(__file__))

def clean_body(text):
    """清理正文中的雜訊"""
    
    # 1. 移除「文字快照：URL」整行（含前後空白）
    text = re.sub(r'文字快照\s*[：:]\s*https?://\S+', '', text)
    
    # 2. 移除媒體 email 署名，如 service@sunmedia.tw (商傳媒 SUN MEDIA)
    text = re.sub(r'\S+@\S+\.\S+\s*\([^)]*\)', '', text)
    # 單獨的 email
    text = re.sub(r'\S+@\S+\.\S{2,}', '', text)
    
    # 3. 移除「・X天前・發表留言」「・X小時前・發表留言」模式
    text = re.sub(r'[\s　]*・\s*\d+\s*(天|小時|分鐘)\s*前\s*・\s*(發表留言|\d+)', '', text)
    # 也移除「媒體名 ・ 時間 ・ 發表留言」
    text = re.sub(r'[\s　]+\S{2,10}\s*・\s*\d+\s*(天|小時|分鐘)\s*前\s*・\s*(發表留言|\d+)', '', text)
    
    # 4. 移除版權/授權聲明（通常在末尾括號內）
    # 如：（本文由 XXX 授權轉載；首圖來源：...）
    text = re.sub(r'[（(]\s*本文由.{2,50}(授權|轉載)[^)）]*[)）]', '', text)
    # 首圖來源
    text = re.sub(r'[（(]\s*首圖來源[：:][^)）]*[)）]', '', text)
    # 圖片來源
    text = re.sub(r'[（(]\s*(圖片|圖)\s*/?\s*來源[：:][^)）]*[)）]', '', text)
    
    # 5. 移除「【看原文連結】」「【原文連結】」
    text = re.sub(r'【[^】]*原文[^】]*】', '', text)
    
    # 6. 清理多餘空行（連續3個以上換行變成2個）
    text = re.sub(r'\n{3,}', '\n\n', text)
    
    # 7. 清理行首行尾空白
    lines = text.split('\n')
    lines = [l.rstrip() for l in lines]
    text = '\n'.join(lines)
    
    return text.strip()

def process_file(txt_path):
    with open(txt_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 分割出 header 和文章
    first_art = content.find('■ [1]')
    if first_art == -1:
        return False
    
    header_part = content[:first_art]
    body_part = content[first_art:]
    
    # 處理每個文章 block
    blocks = body_part.split('----------------------------------------')
    new_blocks = []
    
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        
        title_match = re.search(r'(■\s*\[\d+\]\s*.+)', block)
        if not title_match:
            continue
        
        title_line = title_match.group(1)
        rest = block[title_match.end():].strip()
        
        # 找關鍵字行
        kw_match = re.search(r'(\[關鍵字\]\s*.+)', rest)
        if kw_match:
            kw_line = kw_match.group(1)
            body = rest[kw_match.end():].strip()
            # 移除可能的重複關鍵字行
            body = re.sub(r'^\[關鍵字\]\s*.+\n?', '', body).strip()
        else:
            kw_line = ''
            body = rest
        
        # 清理正文
        body = clean_body(body)
        
        # 重組 block
        parts = [title_line]
        if kw_line:
            parts.append(kw_line)
        parts.append('')
        parts.append(body)
        new_blocks.append('\n'.join(parts))
    
    # 重建檔案
    txt_fname = os.path.basename(txt_path)
    base = txt_fname.replace('_kw_kw.txt', '_kw.txt')
    
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(base + '\n')
        f.write('共 ' + str(len(new_blocks)) + ' 篇文章\n\n')
        for block in new_blocks:
            f.write(block)
            f.write('\n\n----------------------------------------\n\n')
    
    return True

def process_xlsx(xlsx_path, txt_path):
    """從清理後的 txt 重建 xlsx"""
    try:
        from openpyxl import Workbook
        
        with open(txt_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        blocks = content.split('----------------------------------------')
        articles = []
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            title_match = re.search(r'■\s*\[\d+\]\s*(.+)', block)
            if not title_match:
                continue
            title = title_match.group(1).strip()
            kw_match = re.search(r'\[關鍵字\]\s*(.+)', block)
            keywords = kw_match.group(1).strip() if kw_match else ''
            kw_positions = [m.end() for m in re.finditer(r'\[關鍵字\]\s*.+', block)]
            if kw_positions:
                body = block[kw_positions[-1]:].strip()
            else:
                body = block[title_match.end():].strip()
            articles.append((title, keywords, body))
        
        wb = Workbook()
        ws = wb.active
        ws.title = '文章'
        ws.cell(row=1, column=1, value='篇號')
        ws.cell(row=1, column=2, value='標題')
        ws.cell(row=1, column=3, value='自動關鍵字(jieba)')
        ws.cell(row=1, column=4, value='全文內容')
        for idx, (title, kw, body) in enumerate(articles, 1):
            ws.cell(row=idx+1, column=1, value=str(idx))
            ws.cell(row=idx+1, column=2, value=title)
            ws.cell(row=idx+1, column=3, value=kw)
            ws.cell(row=idx+1, column=4, value=body)
        wb.save(xlsx_path)
    except Exception as e:
        print('  [xlsx err] ' + str(e), flush=True)

if __name__ == "__main__":
    try:
        pattern = os.path.join(WORK_DIR, '2026*_kw_kw.txt')
        txt_files = sorted(glob.glob(pattern))
        print('found ' + str(len(txt_files)) + ' files', flush=True)
        
        processed = 0
        for i, txt_path in enumerate(txt_files):
            txt_fname = os.path.basename(txt_path)
            xlsx_path = txt_path.replace('.txt', '.xlsx')
            
            if process_file(txt_path):
                processed += 1
                if os.path.exists(xlsx_path):
                    process_xlsx(xlsx_path, txt_path)
            
            if (i + 1) % 30 == 0:
                print('  ...processed ' + str(i+1) + '/' + str(len(txt_files)), flush=True)
        
        print('', flush=True)
        print('===== DONE =====', flush=True)
        print('  Cleaned: ' + str(processed) + ' files', flush=True)

    except Exception:
        import traceback
        traceback.print_exc()
        sys.stdout.flush()
