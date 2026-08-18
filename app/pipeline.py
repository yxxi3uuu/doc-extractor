# -*- coding: utf-8 -*-
"""
Pipeline: 串接現有腳本的處理流程。
"""

import re
import sys
import traceback
from dataclasses import dataclass, field
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))


@dataclass
class JobStatus:
    job_id: str
    total_files: int
    selected_steps: list[int]
    input_dir: Path
    output_dir: Path
    export_format: str = "both"
    messages: list[str] = field(default_factory=list)
    done: bool = False
    error: str = ""
    stats: dict = field(default_factory=dict)
    # 進度追蹤
    progress_current: int = 0
    progress_total: int = 0


jobs: dict[str, JobStatus] = {}


def log(job: JobStatus, msg: str):
    job.messages.append(msg)


def log_progress(job: JobStatus, current: int, total: int, prefix: str = ""):
    """送出帶百分比的進度"""
    job.progress_current = current
    job.progress_total = total
    pct = int(current / total * 100) if total > 0 else 0
    job.messages.append(f"  {prefix}[{current}/{total}] ({pct}%)")


async def run_pipeline(job_id: str):
    import asyncio

    job = jobs[job_id]
    steps = job.selected_steps
    input_dir = job.input_dir
    output_dir = job.output_dir
    stats = {"extracted": 0, "after_keywords": 0, "filtered_out": 0, "dedup_removed": 0, "final": 0}

    try:
        if 1 in steps:
            log(job, "📄 Step 1: 提取文章中...")
            n = await asyncio.to_thread(_step1_extract, job, input_dir, output_dir)
            stats["extracted"] = n
            log(job, f"✅ Step 1 完成 — 提取出 {n} 篇文章")

        if 2 in steps:
            log(job, "🔑 Step 2: 產生關鍵字中...")
            n = await asyncio.to_thread(_step2_keywords, job, output_dir)
            stats["after_keywords"] = n
            log(job, f"✅ Step 2 完成 — {n} 篇已加關鍵字")

        if 3 in steps:
            log(job, "🗑️ Step 3: 過濾不相關文章...")
            kept, removed = await asyncio.to_thread(_step3_filter, job, output_dir)
            stats["filtered_out"] = removed
            log(job, f"✅ Step 3 完成 — 保留 {kept} 篇，刪除 {removed} 篇")

        if 4 in steps:
            log(job, "🔄 Step 4: 去重複中...")
            kept, removed = await asyncio.to_thread(_step4_dedup, job, output_dir)
            stats["dedup_removed"] = removed
            log(job, f"✅ Step 4 完成 — 保留 {kept} 篇，去重 {removed} 篇")

        if 5 in steps:
            log(job, "🧹 Step 5: 清理正文...")
            await asyncio.to_thread(_step5_clean, job, output_dir)
            log(job, "✅ Step 5 完成")

        if 6 in steps:
            log(job, "📝 Step 6: 整理格式...")
            await asyncio.to_thread(_step6_fix_headers, job, output_dir)
            log(job, "✅ Step 6 完成")

        # 計算最終篇數
        final_dir = _get_latest_txt_dir(output_dir)
        if final_dir:
            total_final = 0
            for f in final_dir.glob("*.txt"):
                total_final += len(re.findall(r'■\s*\[\d+\]', f.read_text(encoding="utf-8")))
            stats["final"] = total_final

        job.stats = stats
        log(job, f"🎉 完成！最終共 {stats['final']} 篇文章")
        job.done = True

    except Exception as e:
        job.error = f"處理失敗: {str(e)}"
        job.messages.append(f"❌ 錯誤: {str(e)}")
        job.done = True
        traceback.print_exc()


# ===========================================================================
# Step 實作
# ===========================================================================

def _step1_extract(job: JobStatus, input_dir: Path, output_dir: Path) -> int:
    """提取文章，回傳總篇數"""
    txt_dir = output_dir / "extracted"
    txt_dir.mkdir(exist_ok=True)

    all_files = sorted(
        f for f in input_dir.iterdir()
        if f.suffix.lower() in (".pdf", ".docx", ".xlsx", ".txt")
    )
    total_articles = 0

    for i, file_path in enumerate(all_files, 1):
        log_progress(job, i, len(all_files), f"提取 {file_path.name} ")
        try:
            ext = file_path.suffix.lower()
            if ext == ".pdf":
                articles = _extract_from_pdf(file_path)
            elif ext == ".docx":
                articles = _extract_from_docx(file_path)
            elif ext == ".xlsx":
                articles = _extract_from_xlsx(file_path)
            elif ext == ".txt":
                articles = _extract_from_txt_file(file_path)
            else:
                continue

            if articles:
                out_path = txt_dir / (file_path.stem + ".txt")
                lines = [file_path.name, f"共 {len(articles)} 篇文章", ""]
                for idx, (title, body) in enumerate(articles, 1):
                    lines.append(f"■ [{idx}] {title}")
                    lines.append("")
                    lines.append(body)
                    lines.append("")
                    lines.append("-" * 40)
                    lines.append("")
                out_path.write_text("\n".join(lines), encoding="utf-8")
                total_articles += len(articles)

        except Exception as e:
            log(job, f"  ⚠️ {file_path.name}: {e}")

    return total_articles


def _extract_from_pdf(pdf_path: Path) -> list[tuple[str, str]]:
    import fitz
    from extract_keyword_context import get_articles, build_article_text

    doc = fitz.open(pdf_path)
    articles_obj = get_articles(doc)
    if articles_obj is None:
        doc.close()
        return []

    results = []
    for article in articles_obj:
        build_article_text(doc, article)
        text = article.text.strip()
        if text:
            results.append((article.title, text))
    doc.close()
    return results


def _extract_from_docx(docx_path: Path) -> list[tuple[str, str]]:
    from docx import Document
    doc = Document(str(docx_path))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    if not paragraphs:
        return []
    title = paragraphs[0][:80]
    body = "\n".join(paragraphs[1:]) if len(paragraphs) > 1 else paragraphs[0]
    return [(title, body)]


def _extract_from_xlsx(xlsx_path: Path) -> list[tuple[str, str]]:
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    results = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        title_col = body_col = None
        if header:
            for idx, h in enumerate(header):
                if h is None:
                    continue
                h_str = str(h).strip().lower()
                if any(k in h_str for k in ("標題", "title", "主題")):
                    title_col = idx
                elif any(k in h_str for k in ("內容", "全文", "content", "body", "正文")):
                    body_col = idx

        if title_col is not None and body_col is not None:
            for row in rows[1:]:
                if row is None:
                    continue
                title = str(row[title_col] or "").strip()
                body = str(row[body_col] or "").strip()
                if title and body:
                    results.append((title, body))
        else:
            all_text = []
            for row in rows:
                if row:
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        all_text.append(" | ".join(cells))
            if all_text:
                results.append((all_text[0][:80], "\n".join(all_text)))

    wb.close()
    return results


def _extract_from_txt_file(txt_path: Path) -> list[tuple[str, str]]:
    content = txt_path.read_text(encoding="utf-8")
    import re as _re
    blocks = _re.split(r"\n-{5,}\n|\n={5,}\n|\n\n\n+", content)
    blocks = [b.strip() for b in blocks if b.strip()]

    if len(blocks) > 1:
        results = []
        for block in blocks:
            lines = block.split("\n")
            title = lines[0][:80] if lines else txt_path.stem
            body = "\n".join(lines[1:]).strip() if len(lines) > 1 else block
            if body:
                results.append((title, body))
        return results
    else:
        lines = content.strip().split("\n")
        title = lines[0][:80] if lines else txt_path.stem
        body = "\n".join(lines[1:]).strip() if len(lines) > 1 else content
        return [(title, body)] if body else []


def _step2_keywords(job: JobStatus, output_dir: Path) -> int:
    """加關鍵字，回傳處理篇數"""
    from add_keywords_fast import parse_txt, extract_auto_keywords, STOPWORDS
    from app.config_manager import load_config

    # 載入使用者自訂停用詞
    config = load_config()
    user_stopwords = set(config.get("stopwords", []))
    combined_stopwords = STOPWORDS | user_stopwords

    txt_dir = output_dir / "extracted"
    kw_dir = output_dir / "with_keywords"
    kw_dir.mkdir(exist_ok=True)

    txt_files = sorted(txt_dir.glob("*.txt"))
    if not txt_files:
        return 0

    total_articles = 0
    for i, f in enumerate(txt_files, 1):
        raw = f.read_text(encoding="utf-8")
        articles_raw = parse_txt(raw)

        # 用合併後的停用詞提取關鍵字
        articles_with_kw = []
        for t, b in articles_raw:
            kws = _extract_keywords_with_stopwords(b, combined_stopwords)
            articles_with_kw.append((t, b, kws))

        _write_kw_txt(kw_dir / (f.stem + "_kw.txt"), f.name, articles_with_kw)
        _write_kw_xlsx(kw_dir / (f.stem + "_kw.xlsx"), f.name, articles_with_kw)
        total_articles += len(articles_with_kw)

        if i % 5 == 0 or i == len(txt_files):
            log_progress(job, i, len(txt_files), "關鍵字 ")

    return total_articles


def _extract_keywords_with_stopwords(text: str, stopwords: set, top_n: int = 10) -> list[str]:
    """用自訂停用詞提取關鍵字"""
    import jieba
    clean = re.sub(r"[【】]", "", text)
    if not clean.strip():
        return []

    word_count = {}
    for word in jieba.cut(clean):
        word = word.strip()
        if len(word) < 2 or word.isdigit() or word in stopwords:
            continue
        if word.endswith("元") or word.endswith("%"):
            continue
        if word[-1] in "及與和或之等的了在為從到並且而是":
            continue
        if word[0] in "及與和或之等的了在為從到並且而是":
            continue
        if "大學" in word:
            continue
        if len(word) > 6:
            continue
        word_count[word] = word_count.get(word, 0) + 1

    sorted_words = sorted(word_count.items(), key=lambda x: x[1], reverse=True)
    return [w for w, _ in sorted_words[:top_n]]


def _write_kw_txt(out_path: Path, source_name: str, articles):
    lines = [source_name, f"共 {len(articles)} 篇文章", ""]
    for idx, (title, body, kws) in enumerate(articles, 1):
        lines.append(f"■ [{idx}] {title}")
        if kws:
            lines.append(f"[關鍵字] {', '.join(kws)}")
        lines.append("")
        lines.append(body)
        lines.append("")
        lines.append("-" * 40)
        lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")


def _write_kw_xlsx(out_path: Path, source_name: str, articles):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "文章"
    ws.append(["篇號", "標題", "自動關鍵字(jieba)", "全文內容"])
    for idx, (title, body, kws) in enumerate(articles, 1):
        ws.append([idx, title, ", ".join(kws), body])
    wb.save(str(out_path))


def _step3_filter(job: JobStatus, output_dir: Path) -> tuple[int, int]:
    """過濾，回傳 (保留數, 刪除數)"""
    from filter_articles import parse_articles, check_ad_by_content, ARTICLE_RE
    from app.config_manager import load_config

    config = load_config()
    user_keywords = config.get("filter_keywords", {})
    settings = config.get("filter_settings", {})

    kw_dir = output_dir / "with_keywords"
    filtered_dir = output_dir / "filtered"
    filtered_dir.mkdir(exist_ok=True)

    txt_files = sorted(kw_dir.glob("*.txt"))
    if not txt_files:
        return (0, 0)

    total_removed = 0
    total_kept = 0

    for fi, f in enumerate(txt_files, 1):
        raw = f.read_text(encoding="utf-8")
        articles = parse_articles(raw)
        kept_sections = []

        for num, title, content in articles:
            if title == "__HEADER__":
                kept_sections.append(content)
                continue

            removed = False
            norm_title = re.sub(r"[\s\u3000]+", "", title)
            title_lower = title.lower()
            norm_lower = norm_title.lower()

            for cat_name, kw_list in user_keywords.items():
                if not settings.get(f"{cat_name}_enabled", True):
                    continue
                need_two = settings.get(f"{cat_name}_need_two", False)
                hits = []
                for kw in kw_list:
                    kw_lower = kw.lower()
                    if kw in title or kw in norm_title or kw_lower in title_lower or kw_lower in norm_lower:
                        hits.append(kw)
                if need_two and len(hits) >= 2:
                    removed = True
                    break
                elif not need_two and hits:
                    removed = True
                    break

            if not removed:
                is_ad, *_ = check_ad_by_content(content)
                if is_ad:
                    removed = True

            if removed:
                total_removed += 1
            else:
                kept_sections.append(content)
                total_kept += 1

        output = ("\n\n" + "-" * 40 + "\n\n").join(kept_sections)
        idx = 0

        def renumber(match):
            nonlocal idx
            idx += 1
            return f"■ [{idx}] {match.group(2).strip()}"

        output = ARTICLE_RE.sub(renumber, output)
        (filtered_dir / f.name).write_text(output, encoding="utf-8")

        if fi % 5 == 0 or fi == len(txt_files):
            log_progress(job, fi, len(txt_files), "過濾 ")

    return (total_kept, total_removed)


def _step4_dedup(job: JobStatus, output_dir: Path) -> tuple[int, int]:
    """去重複，回傳 (保留數, 刪除數)"""
    from dedup_articles import parse_file, find_duplicates, mark_duplicates, write_output

    filtered_dir = output_dir / "filtered"
    deduped_dir = output_dir / "deduped"
    deduped_dir.mkdir(exist_ok=True)

    txt_files = sorted(filtered_dir.glob("*.txt"))
    if not txt_files:
        return (0, 0)

    all_articles = []
    file_order = []
    for f in txt_files:
        arts = parse_file(f)
        all_articles.extend(arts)
        file_order.append(f.name)

    total = sum(1 for a in all_articles if a.title != "__HEADER__")
    log(job, f"  共 {total} 篇文章，比對中...")

    dup_groups = find_duplicates(all_articles)
    mark_duplicates(all_articles, dup_groups)

    removed = sum(1 for a in all_articles if a.title != "__HEADER__" and not a.keep)
    kept = total - removed

    write_output(all_articles, file_order, deduped_dir)
    return (kept, removed)


def _step5_clean(job: JobStatus, output_dir: Path):
    from clean import clean_body

    work_dir = _get_latest_txt_dir(output_dir)
    if not work_dir:
        return

    txt_files = sorted(work_dir.glob("*.txt"))
    for i, txt_path in enumerate(txt_files, 1):
        content = txt_path.read_text(encoding="utf-8")
        first_art = content.find('■ [1]')
        if first_art == -1:
            continue

        body_part = content[first_art:]
        blocks = body_part.split('----------------------------------------')
        new_blocks = []

        for block in blocks:
            block = block.strip()
            if not block:
                continue
            title_match = re.search(r'(■\s*\[\d+\]\s*.+)', block)
            if not title_match:
                continue
            title_line = title_match.group(1)
            rest = block[title_match.end():].strip()

            kw_match = re.search(r'(\[關鍵字\]\s*.+)', rest)
            if kw_match:
                kw_line = kw_match.group(1)
                body = rest[kw_match.end():].strip()
            else:
                kw_line = ''
                body = rest

            body = clean_body(body)
            parts = [title_line]
            if kw_line:
                parts.append(kw_line)
            parts.append('')
            parts.append(body)
            new_blocks.append('\n'.join(parts))

        txt_fname = txt_path.name
        base = txt_fname.replace('_kw_kw.txt', '_kw.txt').replace('_kw.txt', '_kw.txt')
        new_content = base + '\n' + f'共 {len(new_blocks)} 篇文章\n\n'
        for block in new_blocks:
            new_content += block + '\n\n----------------------------------------\n\n'
        txt_path.write_text(new_content, encoding="utf-8")

        if i % 5 == 0 or i == len(txt_files):
            log_progress(job, i, len(txt_files), "清理 ")


def _step6_fix_headers(job: JobStatus, output_dir: Path):
    work_dir = _get_latest_txt_dir(output_dir)
    if not work_dir:
        return

    txt_files = sorted(work_dir.glob("*.txt"))
    for txt_path in txt_files:
        content = txt_path.read_text(encoding="utf-8")
        first_art = content.find('■ [1]')
        if first_art == -1:
            continue
        art_count = len(re.findall(r'■\s*\[\d+\]', content))
        fname = txt_path.name
        base = fname.replace('_kw_kw.txt', '_kw.txt').replace('_kw.txt', '_kw.txt')
        clean_content = base + '\n共 ' + str(art_count) + ' 篇文章\n\n' + content[first_art:]
        if content != clean_content:
            txt_path.write_text(clean_content, encoding="utf-8")

    _rebuild_xlsx(work_dir)
    _copy_final_results(work_dir, output_dir)


def _rebuild_xlsx(work_dir: Path):
    from openpyxl import Workbook
    for txt_path in sorted(work_dir.glob("*.txt")):
        content = txt_path.read_text(encoding="utf-8")
        blocks = content.split('----------------------------------------')
        articles = []
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            title_match = re.search(r'■\s*\[\d+\]\s*(.+)', block)
            if not title_match:
                continue
            title = title_match.group(1).strip()
            kw_match = re.search(r'\[關鍵字\]\s*(.+)', block)
            keywords = kw_match.group(1).strip() if kw_match else ''
            kw_positions = [m.end() for m in re.finditer(r'\[關鍵字\]\s*.+', block)]
            body = block[kw_positions[-1]:].strip() if kw_positions else block[title_match.end():].strip()
            articles.append((title, keywords, body))
        if not articles:
            continue
        xlsx_path = txt_path.with_suffix('.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = '文章'
        ws.append(['篇號', '標題', '自動關鍵字(jieba)', '全文內容'])
        for idx, (title, kw, body) in enumerate(articles, 1):
            ws.append([str(idx), title, kw, body])
        wb.save(str(xlsx_path))


def _copy_final_results(work_dir: Path, output_dir: Path):
    import shutil
    final_dir = output_dir / "final"
    final_dir.mkdir(exist_ok=True)
    for f in work_dir.iterdir():
        if f.suffix in (".txt", ".xlsx", ".csv"):
            shutil.copy2(f, final_dir / f.name)


def _get_latest_txt_dir(output_dir: Path) -> Path | None:
    for dir_name in ["deduped", "filtered", "with_keywords", "extracted"]:
        d = output_dir / dir_name
        if d.exists() and list(d.glob("*.txt")):
            return d
    return None
