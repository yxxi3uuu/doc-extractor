# -*- coding: utf-8 -*-
"""
FastAPI 主入口：文件提取與清理服務
"""

import asyncio
import json
import shutil
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from sse_starlette.sse import EventSourceResponse

from app.pipeline import run_pipeline, JobStatus, jobs
from app import config_manager

app = FastAPI(title="文件提取服務")

# 靜態檔案
STATIC_DIR = Path(__file__).parent / "static"
TEMP_DIR = Path(__file__).parent / "temp"
TEMP_DIR.mkdir(exist_ok=True)

# 歷史紀錄檔
HISTORY_PATH = Path(__file__).parent / "history.json"

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ===========================================================================
# 歷史紀錄管理
# ===========================================================================

def _load_history() -> list[dict]:
    if HISTORY_PATH.exists():
        try:
            return json.loads(HISTORY_PATH.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, IOError):
            return []
    return []


def _save_history(history: list[dict]):
    HISTORY_PATH.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def _add_history(job: JobStatus, stats: dict):
    history = _load_history()
    history.insert(0, {
        "job_id": job.job_id,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "files": job.total_files,
        "steps": job.selected_steps,
        "stats": stats,
    })
    # 只保留最近 50 筆
    history = history[:50]
    _save_history(history)


# ===========================================================================
# 頁面路由
# ===========================================================================

@app.get("/", response_class=HTMLResponse)
async def index():
    return (STATIC_DIR / "index.html").read_text(encoding="utf-8")


@app.get("/settings", response_class=HTMLResponse)
async def settings_page():
    return (STATIC_DIR / "settings.html").read_text(encoding="utf-8")


@app.get("/stopwords", response_class=HTMLResponse)
async def stopwords_page():
    return (STATIC_DIR / "stopwords.html").read_text(encoding="utf-8")


# ===========================================================================
# 上傳與處理 API
# ===========================================================================

@app.post("/api/upload")
async def upload_files(
    files: list[UploadFile] = File(...),
    steps: str = Form("1,2,3,4,5,6"),
    export_format: str = Form("both"),
):
    """
    接收檔案上傳，啟動處理流程。
    steps: 逗號分隔的步驟編號
    export_format: "txt", "xlsx", "both"
    """
    job_id = str(uuid.uuid4())[:8]
    job_dir = TEMP_DIR / job_id
    input_dir = job_dir / "input"
    output_dir = job_dir / "output"
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    ALLOWED_EXT = (".pdf", ".docx", ".xlsx", ".txt")
    saved_files = []
    for f in files:
        ext = "." + f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
        if ext not in ALLOWED_EXT:
            continue
        dest = input_dir / f.filename
        with open(dest, "wb") as buf:
            content = await f.read()
            buf.write(content)
        saved_files.append(dest)

    if not saved_files:
        return {"error": "沒有收到支援的檔案（支援 PDF、Word、Excel、TXT）"}

    selected_steps = [int(s.strip()) for s in steps.split(",") if s.strip().isdigit()]

    jobs[job_id] = JobStatus(
        job_id=job_id,
        total_files=len(saved_files),
        selected_steps=selected_steps,
        input_dir=input_dir,
        output_dir=output_dir,
        export_format=export_format,
    )

    asyncio.create_task(run_pipeline(job_id))

    return {"job_id": job_id, "file_count": len(saved_files), "steps": selected_steps}


@app.get("/api/progress/{job_id}")
async def progress_stream(job_id: str):
    """SSE 即時進度串流"""

    async def event_generator():
        if job_id not in jobs:
            yield {"event": "error", "data": "Job not found"}
            return

        job = jobs[job_id]
        last_msg_idx = 0

        while not job.done:
            while last_msg_idx < len(job.messages):
                msg = job.messages[last_msg_idx]
                yield {"event": "progress", "data": msg}
                last_msg_idx += 1
            await asyncio.sleep(0.3)

        # 送出剩餘訊息
        while last_msg_idx < len(job.messages):
            msg = job.messages[last_msg_idx]
            yield {"event": "progress", "data": msg}
            last_msg_idx += 1

        if job.error:
            yield {"event": "error", "data": job.error}
        else:
            # 儲存歷史紀錄
            _add_history(job, job.stats)
            # 送完成事件 + 統計資訊
            done_data = json.dumps({
                "download_url": f"/api/download/{job_id}",
                "preview_url": f"/api/preview/{job_id}",
                "stats": job.stats,
            }, ensure_ascii=False)
            yield {"event": "done", "data": done_data}

    return EventSourceResponse(event_generator())


@app.get("/api/download/{job_id}")
async def download_result(job_id: str, format: str = ""):
    """下載處理結果 zip"""
    if job_id not in jobs:
        return JSONResponse({"error": "Job not found"}, status_code=404)

    job = jobs[job_id]
    if not job.done or job.error:
        return JSONResponse({"error": "Job not ready"}, status_code=400)

    export_fmt = format or job.export_format or "both"

    zip_path = TEMP_DIR / job_id / f"result_{export_fmt}.zip"
    file_count = 0
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in job.output_dir.rglob("*"):
            if f.is_file() and f.name.startswith("result"):
                continue
            if f.suffix == ".txt" and export_fmt in ("txt", "both"):
                zf.write(f, f.relative_to(job.output_dir))
                file_count += 1
            elif f.suffix == ".xlsx" and export_fmt in ("xlsx", "both"):
                zf.write(f, f.relative_to(job.output_dir))
                file_count += 1
            elif f.suffix == ".csv":
                zf.write(f, f.relative_to(job.output_dir))
                file_count += 1

    if file_count == 0:
        return JSONResponse({"error": "沒有產生任何檔案"}, status_code=400)

    return FileResponse(
        zip_path,
        media_type="application/zip",
        filename=f"result_{job_id}.zip",
    )


@app.get("/api/preview/{job_id}")
async def preview_result(job_id: str):
    """取得處理結果預覽（文章標題 + 關鍵字列表）"""
    if job_id not in jobs:
        return JSONResponse({"error": "Job not found"}, status_code=404)

    job = jobs[job_id]
    if not job.done:
        return JSONResponse({"error": "Job not ready"}, status_code=400)

    import re
    articles = []
    work_dir = _get_latest_txt_dir(job.output_dir)
    if not work_dir:
        return {"articles": []}

    for txt_path in sorted(work_dir.glob("*.txt")):
        content = txt_path.read_text(encoding="utf-8")
        blocks = content.split("----------------------------------------")
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            title_match = re.search(r'■\s*\[\d+\]\s*(.+)', block)
            if not title_match:
                continue
            title = title_match.group(1).strip()
            kw_match = re.search(r'\[關鍵字\]\s*(.+)', block)
            keywords = kw_match.group(1).strip() if kw_match else ""
            # 取前 100 字當預覽
            kw_positions = [m.end() for m in re.finditer(r'\[關鍵字\]\s*.+', block)]
            if kw_positions:
                body = block[kw_positions[-1]:].strip()
            else:
                body = block[title_match.end():].strip()
            preview = body[:150].replace("\n", " ") + ("..." if len(body) > 150 else "")
            articles.append({
                "title": title,
                "keywords": keywords,
                "preview": preview,
                "source": txt_path.name,
            })

    return {"articles": articles, "total": len(articles)}


# ===========================================================================
# 歷史紀錄 API
# ===========================================================================

@app.get("/api/history")
async def get_history():
    """取得處理歷史"""
    history = _load_history()
    # 檢查哪些 job 的檔案還在
    for item in history:
        job_dir = TEMP_DIR / item["job_id"]
        item["available"] = job_dir.exists()
    return history


@app.delete("/api/history/{job_id}")
async def delete_history(job_id: str):
    """刪除單筆歷史"""
    history = _load_history()
    history = [h for h in history if h["job_id"] != job_id]
    _save_history(history)
    # 清理檔案
    job_dir = TEMP_DIR / job_id
    if job_dir.exists():
        shutil.rmtree(job_dir, ignore_errors=True)
    if job_id in jobs:
        del jobs[job_id]
    return {"status": "deleted"}


# ===========================================================================
# 設定 API（過濾關鍵字）
# ===========================================================================

@app.get("/api/config/keywords")
async def get_keywords():
    return config_manager.get_filter_keywords()


@app.post("/api/config/keywords")
async def add_keyword(category: str = Form(...), keyword: str = Form(...)):
    return config_manager.add_keyword(category, keyword)


@app.delete("/api/config/keywords")
async def delete_keyword(category: str, keyword: str):
    return config_manager.remove_keyword(category, keyword)


@app.post("/api/config/category")
async def add_category(category: str = Form(...)):
    return config_manager.add_category(category)


@app.delete("/api/config/category")
async def delete_category(category: str):
    return config_manager.remove_category(category)


# ===========================================================================
# 停用詞 API
# ===========================================================================

@app.get("/api/config/stopwords")
async def get_stopwords():
    config = config_manager.load_config()
    return config.get("stopwords", [])


@app.post("/api/config/stopwords")
async def add_stopword(word: str = Form(...)):
    config = config_manager.load_config()
    stopwords = config.setdefault("stopwords", [])
    if word not in stopwords:
        stopwords.append(word)
        stopwords.sort()
    config_manager.save_config(config)
    return stopwords


@app.delete("/api/config/stopwords")
async def delete_stopword(word: str):
    config = config_manager.load_config()
    stopwords = config.get("stopwords", [])
    if word in stopwords:
        stopwords.remove(word)
    config["stopwords"] = stopwords
    config_manager.save_config(config)
    return stopwords


# ===========================================================================
# 字典檔匯入/管理 API
# ===========================================================================

DICT_DIR = Path(__file__).parent / "dictionaries"
DICT_DIR.mkdir(exist_ok=True)


@app.get("/dictionary", response_class=HTMLResponse)
async def dictionary_page():
    return (STATIC_DIR / "dictionary.html").read_text(encoding="utf-8")


@app.post("/api/dictionary/upload")
async def upload_dictionary(
    file: UploadFile = File(...),
    note: str = Form(""),
):
    """
    上傳字典 Excel 檔，解析後更新 config.json。
    Excel 格式：
    - sheet 'filter_keywords': 欄位「類別」「關鍵字」
    - sheet 'stopwords': 欄位「停用詞」
    - sheet 'settings': 欄位「設定項」「值」
    """
    from openpyxl import load_workbook
    from datetime import datetime

    if not file.filename.lower().endswith(".xlsx"):
        return JSONResponse({"error": "請上傳 .xlsx 格式的 Excel 檔"}, status_code=400)

    # 存檔
    content = await file.read()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    saved_name = f"{timestamp}_{file.filename}"
    saved_path = DICT_DIR / saved_name
    saved_path.write_bytes(content)

    # 解析
    try:
        from io import BytesIO
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)

        config = config_manager.load_config()
        parsed_info = {}

        # 解析 filter_keywords sheet
        if "filter_keywords" in wb.sheetnames:
            ws = wb["filter_keywords"]
            rows = list(ws.iter_rows(values_only=True))
            keywords_dict = {}
            for row in rows[1:]:  # 跳過表頭
                if not row or len(row) < 2:
                    continue
                category = str(row[0] or "").strip()
                keyword = str(row[1] or "").strip()
                if category and keyword:
                    if category not in keywords_dict:
                        keywords_dict[category] = []
                    if keyword not in keywords_dict[category]:
                        keywords_dict[category].append(keyword)
            if keywords_dict:
                config["filter_keywords"] = keywords_dict
                parsed_info["filter_keywords"] = f"{len(keywords_dict)} 個類別，共 {sum(len(v) for v in keywords_dict.values())} 個關鍵字"

        # 解析 stopwords sheet
        if "stopwords" in wb.sheetnames:
            ws = wb["stopwords"]
            rows = list(ws.iter_rows(values_only=True))
            stopwords = []
            for row in rows[1:]:  # 跳過表頭
                if not row or not row[0]:
                    continue
                word = str(row[0]).strip()
                if word and word not in stopwords:
                    stopwords.append(word)
            if stopwords:
                stopwords.sort()
                config["stopwords"] = stopwords
                parsed_info["stopwords"] = f"{len(stopwords)} 個停用詞"

        # 解析 settings sheet
        if "settings" in wb.sheetnames:
            ws = wb["settings"]
            rows = list(ws.iter_rows(values_only=True))
            settings = config.get("filter_settings", {})
            for row in rows[1:]:
                if not row or len(row) < 2:
                    continue
                key = str(row[0] or "").strip()
                val = row[1]
                if key:
                    # 自動轉型
                    if isinstance(val, bool):
                        settings[key] = val
                    elif str(val).lower() in ("true", "1", "是"):
                        settings[key] = True
                    elif str(val).lower() in ("false", "0", "否"):
                        settings[key] = False
                    else:
                        settings[key] = str(val)
            config["filter_settings"] = settings
            parsed_info["settings"] = f"{len(settings)} 項設定"

        wb.close()
        config_manager.save_config(config)

        # 記錄字典歷史
        dict_history = _load_dict_history()
        dict_history.insert(0, {
            "filename": saved_name,
            "original_name": file.filename,
            "time": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "note": note,
            "info": parsed_info,
        })
        dict_history = dict_history[:20]
        _save_dict_history(dict_history)

        return {
            "status": "success",
            "filename": saved_name,
            "parsed": parsed_info,
        }

    except Exception as e:
        return JSONResponse({"error": f"解析失敗: {str(e)}"}, status_code=400)


@app.get("/api/dictionary/history")
async def get_dict_history():
    """取得已上傳的字典歷史"""
    return _load_dict_history()


@app.get("/api/dictionary/current")
async def get_current_dict():
    """取得目前生效的字典內容"""
    config = config_manager.load_config()
    return {
        "filter_keywords": config.get("filter_keywords", {}),
        "stopwords": config.get("stopwords", []),
        "filter_settings": config.get("filter_settings", {}),
    }


@app.get("/api/dictionary/template")
async def download_template():
    """下載字典範本 Excel"""
    from openpyxl import Workbook

    wb = Workbook()

    # filter_keywords sheet
    ws1 = wb.active
    ws1.title = "filter_keywords"
    ws1.append(["類別", "關鍵字"])
    # 填入目前的設定當範例
    config = config_manager.load_config()
    for cat, kws in config.get("filter_keywords", {}).items():
        for kw in kws:
            ws1.append([cat, kw])

    # stopwords sheet
    ws2 = wb.create_sheet("stopwords")
    ws2.append(["停用詞"])
    for w in config.get("stopwords", []):
        ws2.append([w])

    # settings sheet
    ws3 = wb.create_sheet("settings")
    ws3.append(["設定項", "值"])
    for k, v in config.get("filter_settings", {}).items():
        ws3.append([k, str(v)])

    template_path = DICT_DIR / "_template.xlsx"
    wb.save(str(template_path))

    return FileResponse(
        template_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="dictionary_template.xlsx",
    )


DICT_HISTORY_PATH = Path(__file__).parent / "dict_history.json"


def _load_dict_history() -> list[dict]:
    if DICT_HISTORY_PATH.exists():
        try:
            return json.loads(DICT_HISTORY_PATH.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, IOError):
            return []
    return []


def _save_dict_history(history: list[dict]):
    DICT_HISTORY_PATH.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


# ===========================================================================
# 工具函式
# ===========================================================================

def _get_latest_txt_dir(output_dir: Path) -> Path | None:
    for dir_name in ["deduped", "filtered", "with_keywords", "extracted"]:
        d = output_dir / dir_name
        if d.exists() and list(d.glob("*.txt")):
            return d
    return None
